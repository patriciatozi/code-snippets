# 1. CREATE THE FILE

from openpyxl import Workbook

wb = Workbook()

# Work on the active worksheet
ws = wb.active

# Define a sheet's name
ws.title = 'Pessoal'

# Data can be assigned directly to cells
ws['A1'] = 'Nome'
ws['B1'] = 'Idade'

# Rows can also be appended
ws.append(['João', 48])

# Save the file
wb.save('sample.xlsx')